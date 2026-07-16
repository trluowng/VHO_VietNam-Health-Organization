"""Sync first_aid_ground_truth.xlsx into rules.seed.json.

Reads the ground-truth spreadsheet and converts every row whose ID is not yet
present in rules.seed.json into the JSON rule schema, then appends only those
new rules. Rows already covered by an existing rule (matched by the numeric
ID prefix, e.g. "FA-006" or "GT-042" -- regardless of any descriptive slug
appended to the JSON rule's full id) are left untouched, so re-running this
script never creates duplicates.

Usage:
    python scripts/sync_ground_truth_to_rules.py
    python scripts/sync_ground_truth_to_rules.py --dry-run
"""
from __future__ import annotations

import argparse
import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data" / "first_aid"
DEFAULT_XLSX = DATA_DIR / "first_aid_ground_truth.xlsx"
DEFAULT_RULES = DATA_DIR / "rules.seed.json"
SHEET_NAME = "First-aid Ground Truth"

ID_PREFIX_RE = re.compile(r"^(FA|GT)-(\d+)")

# Maps the free-text "Source" cell (one org per "; "-separated segment) to the
# source_id entries curated in source_registry.json. Extend this alongside any
# new source added there.
SOURCE_NAME_TO_ID = {
    "WHO-ICRC Basic Emergency Care (2018)": "who_icrc_basic_emergency_care_2018",
    "WHO - Prehospital emergency care: pocket reference": "who_prehospital_emergency_care_2026",
    "WHO (who.int)": "who_general",
    "NHS - First aid (nhs.uk)": "nhs_first_aid_overview",
    "CDC (cdc.gov)": "cdc_general",
    "Mayo Clinic Patient Care": "mayo_clinic_patient_care",
    "MedlinePlus - U.S. National Library of Medicine": "medlineplus_general",
    "American Red Cross - First Aid": "american_red_cross_bleeding",
    "Bộ Y tế Việt Nam / Cục Phòng bệnh (moh.gov.vn, vncdc.gov.vn)": "bo_y_te_vietnam",
}

FALLBACK_SOURCE_ID = "bo_y_te_vietnam"


def canonical_id(raw_id: str) -> str:
    """Reduce any rule/row id to its 'FA-006' / 'GT-042' canonical prefix."""
    match = ID_PREFIX_RE.match(raw_id.strip())
    return f"{match.group(1)}-{match.group(2)}" if match else raw_id.strip()


def strip_diacritics(text: str) -> str:
    normalized = unicodedata.normalize("NFD", text)
    without_marks = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
    return without_marks.replace("Đ", "D").replace("đ", "d")


def slugify(text: str, sep: str = "_", max_len: int = 60) -> str:
    ascii_text = strip_diacritics(text).lower()
    ascii_text = re.sub(r"[^a-z0-9]+", sep, ascii_text).strip(sep)
    return ascii_text[:max_len].strip(sep)


def split_signals(cell_text: str) -> list[str]:
    """'Dau hieu nhan biet' cells use ',' between short symptom phrases."""
    parts = re.split(r",\s*", (cell_text or "").strip())
    return [p.strip() for p in parts if p.strip()]


def split_actions(cell_text: str) -> list[str]:
    """'Huong dan so cuu' / 'Khong nen lam' cells use ';' between items -- a
    plain ',' inside one of those items (e.g. a parenthetical aside) must not
    be split further."""
    parts = re.split(r";\s*", (cell_text or "").strip())
    return [p.strip() for p in parts if p.strip()]


def classify_decision(urgency_text: str) -> str:
    return "emergency" if (urgency_text or "").strip().startswith("Khẩn cấp") else "non_emergency"


def source_ids_for(source_cell: str) -> list[str]:
    names = [n.strip() for n in (source_cell or "").split(";")]
    ids: list[str] = []
    for name in names:
        source_id = SOURCE_NAME_TO_ID.get(name)
        if source_id and source_id not in ids:
            ids.append(source_id)
    return ids


def row_to_rule(raw_id: str, name_vi: str, category: str, urgency: str, signals_vi: str,
                 actions_vi: str, never_do_vi: str, escalation_vi: str, source_cell: str) -> dict:
    signal_phrases = split_signals(signals_vi) or [name_vi]
    action_phrases = split_actions(actions_vi)
    never_do_phrases = split_actions(never_do_vi)
    category_slug = slugify(category, sep="_")

    evidence = [{"source_id": sid, "topic": category_slug} for sid in source_ids_for(source_cell)]
    if not evidence:
        evidence = [{"source_id": FALLBACK_SOURCE_ID, "topic": category_slug}]

    return {
        "id": f"{raw_id}-{slugify(name_vi, sep='-', max_len=40)}",
        "status": "needs_clinician_review",
        "scope": "adult_or_child",
        "signals_any": [slugify(p) for p in signal_phrases if slugify(p)],
        "signals_any_vi": signal_phrases,
        "decision": classify_decision(urgency),
        "education_actions": [slugify(a) for a in action_phrases if slugify(a)],
        "education_actions_vi": action_phrases,
        "never_do": [slugify(n) for n in never_do_phrases if slugify(n)],
        "never_do_vi": never_do_phrases,
        "escalation_vi": escalation_vi,
        "evidence": evidence,
        "review": {"clinical_approver": None, "reviewed_at": None, "expires_at": None},
        "name_vi": name_vi,
        "category_vi": category,
    }


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Sync first_aid_ground_truth.xlsx into rules.seed.json (idempotent, skips existing ids)"
    )
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--rules", type=Path, default=DEFAULT_RULES)
    parser.add_argument("--dry-run", action="store_true", help="Preview counts without writing rules.seed.json")
    args = parser.parse_args()

    wb = load_workbook(args.xlsx, data_only=True)
    ws = wb[SHEET_NAME]

    with args.rules.open("r", encoding="utf-8") as fh:
        rules_doc = json.load(fh)

    existing_canonical_ids = {canonical_id(rule["id"]) for rule in rules_doc.get("rules", [])}

    added = []
    skipped_existing = []
    skipped_blank = 0
    total_rows = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] is None:
            skipped_blank += 1
            continue
        total_rows += 1
        (_stt, raw_id, name_vi, category, urgency, signals_vi, actions_vi,
         never_do_vi, escalation_vi, _status_vi, source_cell, _phan_loai) = row[:12]

        canon = canonical_id(str(raw_id))
        if canon in existing_canonical_ids:
            skipped_existing.append(str(raw_id))
            continue

        rule = row_to_rule(
            str(raw_id), str(name_vi), str(category), str(urgency), str(signals_vi),
            str(actions_vi), str(never_do_vi), str(escalation_vi), str(source_cell),
        )
        added.append(rule)
        existing_canonical_ids.add(canon)

    if not args.dry_run and added:
        rules_doc.setdefault("rules", []).extend(added)
        with args.rules.open("w", encoding="utf-8") as fh:
            json.dump(rules_doc, fh, ensure_ascii=False, indent=2)
            fh.write("\n")

    print(f"Ground-truth rows read: {total_rows} (blank rows skipped: {skipped_blank})")
    print(f"Already in rules.seed.json (skipped, no duplicates created): {len(skipped_existing)}")
    suffix = " (dry-run, rules.seed.json not written)" if args.dry_run else ""
    print(f"Newly added rules: {len(added)}{suffix}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
